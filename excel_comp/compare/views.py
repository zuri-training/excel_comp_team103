from django.shortcuts import render
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

# Create your views here.
def unauth(request):
    return render(request, "demo.html")


def upload(request):
    if request.method == 'POST':
        excel_file = request.FILES["file"]
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        worksheet = wb[sheets[0]]
        excel_data = list()
        unique_rows = set()
        duplicate_rows = list()

        for row in worksheet.iter_rows():
            row_data = list()

            for cell in row:
                row_data.append(str(cell.value))
                
            if row_data[0] in unique_rows:
                print("Got here")
                duplicate_rows.append(row_data[0])
                print(duplicate_rows)
            unique_rows.add(row_data[0])
            excel_data.append(row_data)
        create_excel_with_data(excel_data, duplicate_rows)
    return render(request, 'demo.html')


def create_excel_with_data(data, duplicates):
    wb = openpyxl.Workbook()
    ws = wb.active

    for d in data:
        ws.append(d)
    red_fill =  PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(fill=red_fill)
    r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    formula_array = []

    for fa in duplicates:
        formula_array.append(f'$A1="{fa}"')
    r.formula = formula_array
    ws.conditional_formatting.add("A1:O100", r)
    wb.save("uploads/sample.xlsx")